# -------------------------------
# Section 1: Imports and Constants
# -------------------------------

# Built-in libraries
import re         # Regular expressions for ID validation
import math       # Used for ceiling division of chunks
import argparse   # CLI argument parsing
from pathlib import Path  # Cross-platform file and directory handling

# openpyxl for Excel operations
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from copy import copy  # Deep copy for cell style objects

# Performance tracking
import time
start_time = time.time()

# Regular expression to validate 12-digit numeric product IDs
ID_REGEX = re.compile(r"^\d{12}$")

# Default prefix for output files
DEFAULT_OUTPUT_PREFIX = "split"

# ---------------------------------
# Argument parser configuration
# ---------------------------------

def parse_args():
    """
    Parse and return command-line arguments.
    This sets up flexible CLI control over input/output paths, 
    splitting strategy, and optional dry-run simulation.
    """
    parser = argparse.ArgumentParser(
        description="Split a large Excel file into smaller files with preserved styled headers."
    )
    
    parser.add_argument(
        "-i", "--input", 
        type=Path, 
        required=True, 
        help="Path to the input Excel file (.xlsx)"
    )
    parser.add_argument(
        "-o", "--output-dir", 
        type=Path, 
        default=Path.cwd(), 
        help="Directory where output files will be saved"
    )
    
    # Only one of these two options must be specified
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "-p", "--per-file", 
        type=int, 
        help="Number of entries per output file"
    )
    group.add_argument(
        "-n", "--num-files", 
        type=int, 
        help="Number of output files to create"
    )

    parser.add_argument(
        "-s", "--sheet-name", 
        type=str, 
        default=None, 
        help="Name of the worksheet to process (default: first visible sheet)"
    )
    parser.add_argument(
        "--prefix", 
        type=str, 
        default=DEFAULT_OUTPUT_PREFIX, 
        help="Prefix for the output file names"
    )
    parser.add_argument(
        "--dry-run", 
        action="store_true", 
        help="Simulate the operation and report number of files without writing output"
    )
    
    return parser.parse_args()

# -------------------------------
# Section 2: Load and Validate Input File
# -------------------------------

def load_workbook_and_sheet(input_path: Path, sheet_name: str | None = None) -> Worksheet:
    """
    Load the Excel workbook and return the specified worksheet.

    Parameters:
        input_path (Path): Path to the input Excel file.
        sheet_name (str | None): Optional name of the worksheet to load. 
                                 If not provided, the second sheet in the workbook is used.

    Returns:
        Worksheet: The selected worksheet object.

    Raises:
        RuntimeError: If the file cannot be loaded.
        ValueError: If the specified sheet name is not found or if the sheet appears empty.
    """
    try:
        wb = load_workbook(input_path, data_only=False)
    except Exception as e:
        raise RuntimeError(f"Failed to load workbook: {e}")

    if sheet_name:
        # Use the provided sheet name if it exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        ws = wb[sheet_name]
    else:
        # Default to the second sheet (index 1), assuming the first may be hidden
        ws = wb[wb.sheetnames[1]]

    if ws.max_row <= 1:
        raise ValueError("Worksheet appears to be empty or lacks data beyond the header.")

    return ws

# -------------------------------
# Section 3: Extract and Filter Valid Rows
# -------------------------------

def detect_data_start_row(ws: Worksheet, id_column_index: int = 5) -> int:
    """
    Detect the first row containing a valid 12-digit product ID.

    Parameters:
        ws (Worksheet): The worksheet to scan.
        id_column_index (int): Zero-based index of the column containing the ID (default: 5 → column F).

    Returns:
        int: The row number (1-based) where valid data starts.

    Raises:
        ValueError: If no valid ID is found in the specified column.
    """
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell = row[id_column_index]
        if cell.value and isinstance(cell.value, (int, str)):
            id_str = str(cell.value).strip()
            if ID_REGEX.match(id_str):
                return cell.row
    raise ValueError("No valid data rows with 12-digit ID found.")


def extract_header(ws: Worksheet, data_start_row: int) -> list[list[Cell]]:
    """
    Extract the header rows above the detected data row.

    Parameters:
        ws (Worksheet): The worksheet object.
        data_start_row (int): Row number where data begins.

    Returns:
        list[list[Cell]]: A list of header rows, each row being a list of Cell objects.
    """
    header_rows = []
    for row in ws.iter_rows(min_row=1, max_row=data_start_row - 1):
        header_rows.append([cell for cell in row])
    return header_rows


def extract_valid_data_rows(ws: Worksheet, data_start_row: int, id_column_index: int = 5) -> list[list[Cell]]:
    """
    Extract valid data rows starting from a given row where the ID column contains a 12-digit number.

    Parameters:
        ws (Worksheet): The worksheet to extract from.
        data_start_row (int): The first row to check for data.
        id_column_index (int): Zero-based index of the ID column (default: 5 → column F).

    Returns:
        list[list[Cell.value]]: List of row values (not Cell objects) that match the 12-digit ID rule.
    """
    valid_rows = []
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
        cell = row[id_column_index]
        if cell.value and isinstance(cell.value, (int, str)):
            id_str = str(cell.value).strip()
            if ID_REGEX.match(id_str):
                valid_rows.append([cell.value for cell in row])  # Only values, not Cell objects
    return valid_rows

# -------------------------------
# Section 4: Split Rows into Chunks
# -------------------------------

def split_rows_into_chunks(rows: list[list[Cell]], per_file: int | None = None, num_files: int | None = None) -> list[list[list[Cell]]]:
    """
    Split a list of data rows into evenly sized chunks for output files.

    Parameters:
        rows (list[list[Cell]]): The list of data rows to split.
        per_file (int | None): Desired number of entries per output file (mutually exclusive with num_files).
        num_files (int | None): Desired number of output files (mutually exclusive with per_file).

    Returns:
        list[list[list[Cell]]]: A list of row chunks. Each chunk is a list of rows.

    Raises:
        ValueError: If neither per_file nor num_files is specified.
    """
    total_rows = len(rows)

    if per_file is not None:
        chunk_size = per_file
    elif num_files is not None:
        chunk_size = math.ceil(total_rows / num_files)
    else:
        raise ValueError("Either per_file or num_files must be specified.")

    # Split the full list into consecutive chunks of chunk_size
    chunks = [
        rows[i:i + chunk_size]
        for i in range(0, total_rows, chunk_size)
    ]

    return chunks

# -------------------------------
# Section 5: Write Output Files
# -------------------------------

def copy_cell(source: Cell, target: Cell):
    """
    Copy the value and style from a source cell to a target cell.

    Parameters:
        source (Cell): The original cell with content and styling.
        target (Cell): The target cell in the new workbook.
    """
    if source.has_style:
        target.value = source.value
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.number_format = source.number_format  # Safe to copy directly (immutable)
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)


def write_chunks_to_files(
    header_rows: list[list[Cell]],
    row_chunks: list[list[list[Cell]]],
    output_dir: Path,
    merged_ranges, 
    column_widths,
    row_height,
    prefix: str = "split"
):
    """
    Write each chunk of data into a separate Excel file, preserving header styles and layout.

    Parameters:
        header_rows (list[list[Cell]]): Header rows to be copied with formatting.
        row_chunks (list[list[list[Cell]]]): List of data row chunks to write.
        output_dir (Path): Directory where output files will be saved.
        merged_ranges (list[str]): Merged cell ranges to be applied to the header.
        column_widths (dict[str, float]): Mapping of column letters to their widths.
        row_height (float): Specific row height to apply (used for row 2).
        prefix (str): Prefix for output file names.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    for i, chunk in enumerate(row_chunks, start=1):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Write header with styles
        for r_idx, row in enumerate(header_rows, start=1):
            for c_idx, cell in enumerate(row, start=1):
                new_cell = ws.cell(row=r_idx, column=c_idx)
                copy_cell(cell, new_cell)

        # Re-apply merged cells from original header
        for rng in merged_ranges:
            ws.merge_cells(rng)

        # Re-apply column widths (with M fallback)
        for col_letter, width in column_widths.items():
            if width:
                ws.column_dimensions[col_letter].width = width
                if col_letter == "M":
                    ws.column_dimensions[col_letter].width = column_widths["L"]

        # Apply custom row height (row 2 only)
        ws.row_dimensions[2].height = row_height

        # Write data rows (only values, not styles)
        for r_offset, row in enumerate(chunk, start=len(header_rows) + 1):
            for c_idx, cell in enumerate(row, start=1):
                new_cell = ws.cell(row=r_offset, column=c_idx)
                new_cell.value = row[c_idx - 1]

        # Hide unused rows and columns (to match original layout)
        ws.row_dimensions[1].hidden = True
        ws.row_dimensions[5].hidden = True
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].hidden = True

        # Save the file
        output_file = output_dir / f"{prefix}_{i:02}.xlsx"
        wb.save(output_file)
        print(f"Written: {output_file}")

# -------------------------------
# Section 6: Main Logic
# -------------------------------

def main():
    """
    Main execution function. Orchestrates the file splitting workflow:
    - Parses CLI arguments
    - Loads the workbook and worksheet
    - Detects header and data boundaries
    - Splits valid data rows into chunks
    - Writes output files with preserved formatting
    - Optionally simulates the process with a dry-run
    """
    args = parse_args()

    # Load the specified worksheet (or second sheet by default)
    ws = load_workbook_and_sheet(args.input, args.sheet_name)

    # Extract metadata from the worksheet
    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
    column_widths = {
        get_column_letter(i): ws.column_dimensions[get_column_letter(i)].width
        for i in range(1, ws.max_column + 1) 
    }
    row_height = ws.row_dimensions[2].height

    # Identify start of data by finding first valid 12-digit ID
    id_column = 5  # Column F (0-based index)
    data_start_row = detect_data_start_row(ws=ws, id_column_index=id_column)

    # Extract header and valid data rows
    header_rows = extract_header(ws=ws, data_start_row=data_start_row)
    valid_data = extract_valid_data_rows(ws, data_start_row=data_start_row, id_column_index=id_column)

    # Split valid rows into chunks based on --per-file or --num-files
    chunks = split_rows_into_chunks(
        valid_data,
        per_file=args.per_file,
        num_files=args.num_files
    )

    # Optional: Simulate result without writing files
    if args.dry_run:
        print(f"[Dry Run] Total valid entries: {len(valid_data)}")
        print(f"[Dry Run] Number of files to be created: {len(chunks)}")
        return

    # Generate Excel files with preserved layout and header
    write_chunks_to_files(
        header_rows=header_rows,
        row_chunks=chunks,
        output_dir=args.output_dir,
        merged_ranges=merged_ranges,
        column_widths=column_widths,
        row_height=row_height,
        prefix=args.prefix
    )

    # Print total execution time
    print("--- %s seconds ---" % (time.time() - start_time))


# Entry point for script execution
if __name__ == "__main__":
    main()